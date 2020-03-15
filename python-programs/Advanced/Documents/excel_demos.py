import openpyxl
import os

print(os.getcwd())
os.chdir(r'D:\Knowledge\Python\git\python-programs\Advanced\Documents')
print(os.getcwd())

workbook = openpyxl.open('example.xlsx', read_only=False)
print(workbook.sheetnames)
sheet = workbook['Sheet1']
cell = sheet['A1']
print(str(cell.value))
workbook.close()

wb = openpyxl.Workbook()
wb['Sheet'].title = 'changed_sheet_name'
wb.create_sheet('my_sheet')
my_sheet = wb['my_sheet']
my_sheet['A1'].value = 5
my_sheet['B1'].value = 'Ravi Jiyani'
my_sheet['C1'].value = '01/01/1991'
wb.save('my_workbook.xlsx')

wb.close()